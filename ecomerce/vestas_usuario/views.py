from django.shortcuts import render, redirect
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import  PatternFill, Border, Side, Alignment, Protection, Font

from vestas_usuario.models import Producto , Order , date_capana
from django.views.generic import  View ,TemplateView , ListView , UpdateView ,CreateView , DeleteView ,DetailView
from django.urls import reverse_lazy

from django.conf import settings
from django.core.mail import  EmailMultiAlternatives 

from usuarios.models import  request_of_product , Profile

from .forms import ProductForm , Pago , campana , reporte

import time

from  .envio_whatsapp import envio  , datos , envio_2
import datetime
from django.contrib import messages
from datetime import date
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import webbrowser 

from django.db.models import Q




from django.template.loader import get_template , render_to_string





class factura (TemplateView):

	model= Order
	template_name= 'sesion/factura.html'
	
	def get_context_data(self , **kwargs):
		context= {}
		dato = self.model.objects.all()
		longitud = len(dato)
		ultimo = dato[longitud-1]
		envio = ultimo
		context['object'] = envio
		return context


	def get(self , request , *args , **kwargs):

		return render(request , self.template_name , self.get_context_data())
	
	
def ListCatalogo(request):
	queryset= request.GET.get("buscar")
	catalogo = Producto.objects.all()
	if queryset:
		catalogo = Producto.objects.filter(
			Q(title__icontains = queryset)|
			Q(description__icontains = queryset)
			).distinct()
	
		
	return render(request , 'usuario/listacatalogo.html' , {'catalogo':catalogo})



"""class create_marketin(ListView):
	model= date_capana
	template_name='usuario/listproduc_marketin_realizada.html'
	queryset = date_capana.objects.all()
	context_object_name ='campana'"""





class contacto(TemplateView):

	template_name='usuario/contacto.html'



class list_catalogo(ListView):
	paginate_by = 9

	model=Producto
	template_name = 'usuario/listacatalogo.html'
	queryset = Producto.objects.all()
	context_object_name ='catalogo'





class list_article(ListView):
	paginate_by = 9
	model=Producto
	template_name = 'usuario/listproduc.html'
	
	queryset = Producto.objects.filter(categoria=1).order_by('id')
	print(queryset)
	context_object_name ='producto'



	


	




class Create_Product(CreateView):  


	model= Producto
	template_name ='usuario/crear_producto.html'
	form_class= ProductForm
	
	success_url= reverse_lazy('ventas:articulos')


	def get_context_data( self , **kwargs):
		

		context = super().get_context_data(** kwargs)
		context['form2']= ProductForm
		return context











def  prequest_of_product(request):  



	if request.method =="POST":


		#send_mail(subjects , message , email_from , RECIPIENT_LIST)

		dato = request_of_product()
		dato.user = request.user
		dato.descripcion = request.POST["mensaje"] 
		dato.save()
		return render(request , "usuario/listproduc.html")
	
	
	return render(request, "usuario/contacto.html")



class update_product(UpdateView):
	model= Producto
	fields=['title',
			'price',
			'discount_price',
			'description',
			'image',
			'image2',
			'image3',
			'image4',
			'detalle',
			'detalle2',
			'detalle3',
			'detalle4',
			'categoria',
			'disponible',

	]
	from_clas=ProductForm
	template_name='usuario/modal_editar_producto.html'
	success_url= reverse_lazy('ventas:articulos')


class list_request_of_product(ListView):
	model = request_of_product
	template_name = 'usuario/list_request_of_producto.html'
	queryset = model.objects.all()
	context_object_name = 'dato'

	
	
	

class delete_product(DeleteView):

	model= Producto 
	template_name= 'usuario/delete_product.html'
	success_url= reverse_lazy('ventas:articulos')



class politica_seguridad(TemplateView):
	template_name ='sesion/politica_de_seguridad.html'


class detail_product(DetailView):

	paginate_by = 9
	model = Producto
	template_name='usuario/Productdetail.html'
	pk_url_kwarg = 'pk'
	queryset= Producto.objects.all()

	def get_context_data(self , ** kwargs):
		context = super().get_context_data(**kwargs)
    	
		context['info']= Producto.objects.all()

		return context



def seng_message_whatspp(request):


	
	print("si hay solicitud")


	fecha_actual = date.today()

	sal = envio(request)
	
	sal.send_message()
	

	
	dato = Order.objects.get(user=request.user, ordered=False)
	dato.ordered = True
	dato.ordered_date = fecha_actual
	dato.save()

	
	# I will send email with date of pay 

	"""try:
		mailseerver= smtplib.SMTP(settings.EMAIL_HOST , settings.EMAIL_PORT)
		mailseerver.starttls()
		mailseerver.ehlo()
		mailseerver.login(settings.EMAIL_HOST_USER , settings.EMAIL_HOST_PASSWORD)


		mensaje = MIMEMultipart()
		mensaje['From'] = settings.EMAIL_HOST_USER
		mensaje['To'] = 'coroemma@hotmail.com'
		mensaje['Subject']='tienes un correo'


		content = render_to_string('sesion/factura.html' , {'object': dato})
		mensaje.attach(MIMEText(content , 'html'))
		mailseerver.sendmail(settings.EMAIL_HOST_USER , 'coroemma@hotmail.com' , mensaje.as_string())
	except Exception as e:
		print(e)
	
	return redirect('ventas:factura')"""	
	
	
	return redirect('ventas:orden_realizada') 



def seng_message_whatspp_2(request , id_dato):

	

	sal = envio_2(request )
	
	sal.send_message(request , id_dato)
	
	
	
	return redirect('ventas:orden_realizada')

	



class OrdenrealizadaAdministrativo(View):


	def get(self, *args, **kwargs):
       
		order = Order.objects.filter( ordered=True).order_by('-ordered_date')
		context = {
			'object': order
		}

		return render(self.request, 'carro/order_summary_realizada_administrador.html', context)

    

    
    	



class Ordenrealizada(View):


    def get(self, *args, **kwargs):
       
        order = Order.objects.filter(user=self.request.user, ordered=True).order_by('-ordered_date')
        context = {
             'object': order
            }
        
        return render(self.request, 'carro/order_summary_realizada_user.html', context)
     	


class detalle_compra_solicitud(DetailView):

	model = request_of_product
	template_name='carro/solicitud_compra_detail.html'
	pk_url_kwarg = 'pk'
	#queryset= Order.objects.filter


class detalle_compra(DetailView):

	model = Order
	template_name='carro/compra_detail.html'
	pk_url_kwarg = 'pk'
	#queryset= Order.objects.filter


class update_Pago(UpdateView):
	model= Order
	fields=['pago_realizado'
			

	]
	
	
	template_name='carro/modal_editar_pago.html'

	success_url= reverse_lazy('ventas:orden_administrador')


class modal_reporte(View):
	model = reporte

	def get_context_data(self , **kwargs):
		context={}
		context['form']= self.model
		print(context)
		return context

	def get( self , request , *args , ** kwargs):

		print(self.get_context_data())
	
		return render(request , 'carro/modal_reporte.html' , self.get_context_data())	

	




class reporte_excel(TemplateView):

	def get(self , request , *args , **kwargs):
		
		print(request.GET)
		
		dato = request.GET.get('fecha')
		dato2 = request.GET.get('fecha_final')#date__range=[startdate, enddate]
		#query = Order.objects.all()


		query = Order.objects.filter(ordered_date__range=[dato, dato2])

		if not dato and not dato2:
			query = Order.objects.all()

		wb= Workbook()
		ws = wb.active
		#bandera = True
		
		controlador = 4
		
		
			#print(sheet.colum)
		ws['B1'].alignment = Alignment(horizontal = "center",vertical = "center")
		ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )

		ws['B1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
		ws['B1'].font = Font(name = 'Calibri', size = 12, bold = True)

		ws['B1'] = 'REPORTE DE VENTAS MULTIMOTORES GEMCA'

		ws.merge_cells('B1:E1')

		ws.row_dimensions[1].height = 25

		ws.column_dimensions['B'].width = 20
		ws.column_dimensions['C'].width = 20
		ws.column_dimensions['D'].width = 20
		ws.column_dimensions['E'].width = 20
		ws.column_dimensions['F'].width = 20
		ws.column_dimensions['G'].width = 20
		ws.column_dimensions['H'].width = 20
		ws.column_dimensions['I'].width = 20
		ws.column_dimensions['J'].width = 20

		ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
		ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['B3'] = 'Nombres'

		ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
		ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['C3'] = 'Apellidos'



		ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
		ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['D3'] = 'Telefono'
		ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
		ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['E3'] = 'Email'

		ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['F3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
		ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['F3'] = 'Tipo de Moto'

		ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['G3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
		ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['G3'] = 'Estado'

		ws['H3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['H3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
		ws['H3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['H3'] = 'Nombre Prod.'


		ws['I3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['I3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
		ws['I3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['I3'] = 'Cantidad'



		ws['J3'].alignment = Alignment(horizontal = "center", vertical = "center")
		ws['J3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
		ws['J3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
		ws['J3'].font = Font(name = 'Calibro', size = 10, bold = True)
		ws['J3'] = 'Total'

		cont= 1
		for q in query:
			#Pintamos los datos en el reporte
			ws.cell(row = controlador , column = 2).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador , column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 2).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 2).value = q.user.first_name

			ws.cell(row = controlador, column = 3).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
																	top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 3).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 3).value = q.user.last_name

			ws.cell(row = controlador, column = 4).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 4).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 4).value = q.user.Telefono


			ws.cell(row = controlador, column = 5).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 5).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 5).value = q.user.email




			ws.cell(row = controlador, column = 6).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 6).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 6).value = q.user.profile.interest.category_name


			ws.cell(row = controlador, column = 7).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 7).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 7).value = q.user.profile.state
			total = 0
			nombre = ','
			cantidad = 0
			for dato in q.roducto.all():
				nombre = nombre + ', ' + dato.producto.title

				cantidad += dato.quantity
				
				total = dato.get_total_item_price()

			ws.cell(row = controlador, column = 8).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 8).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 8).value = nombre

			ws.cell(row = controlador, column = 9).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 9).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 9).value = cantidad


			ws.cell(row = controlador, column = 10).alignment = Alignment(horizontal = "center")
			ws.cell(row = controlador, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
			ws.cell(row = controlador, column = 10).font = Font(name = 'Calibri', size = 8)
			ws.cell(row = controlador, column = 10).value = total


			controlador +=1
			print(cont)

		#establecer nombre de archivo

		nombre_archivo ="reportdeVentas.xlsx"

		response = HttpResponse(content_type= 'application/ms-excel')
		contenido = 'attachment; filename = {0}'.format(nombre_archivo)
		response['Content-Disposition'] = contenido
		wb.save(response)
		return response	


class index_email(View):
	model= date_capana
	template_name= 'sesion/email3.html'
	
	def get_context_data(self , **kwargs):
		context= {}
		dato = date_capana.objects.all()
		longitud = len(dato)
		ultimo = dato[longitud-1]
		envio = ultimo
		context['datos'] = envio
		return context


	def get(self , request , *args , **kwargs):

		return render(request , self.template_name , self.get_context_data())			
								

def create_email( email , subject , template_path  , context ):
	template = get_template(template_path)
	content = template.render(context)
	mail = EmailMultiAlternatives(
		subject= subject,
		body='',
		from_email= settings.EMAIL_HOST_USER,
		to =[
			email
		]


		)
	mail.attach_alternative(content , 'text/html/img')
	return mail



		

def send_email(request):

	send_welcome = create_email(
		'coroemma@hotmail.com',
		'live en youtube',
		'sesion/email.html',
		{'username':'angelo2611'}
		)


	send_welcome.send(fail_silently = False)

	return redirect('ventas:index')



def try_send_mail(dato):
	
	try:
		mailseerver= smtplib.SMTP(settings.EMAIL_HOST , settings.EMAIL_PORT)
		mailseerver.starttls()
		mailseerver.ehlo()
		mailseerver.login(settings.EMAIL_HOST_USER , settings.EMAIL_HOST_PASSWORD)


		mensaje = MIMEMultipart()
		mensaje['From'] = settings.EMAIL_HOST_USER
		mensaje['To'] = 'coroemma@hotmail.com'
		mensaje['Subject']='tienes un correo'


		content = render_to_string('sesion/email3.html' , dato)
		mensaje.attach(MIMEText(content , 'html'))
		mailseerver.sendmail(settings.EMAIL_HOST_USER , 'coroemma@hotmail.com' , mensaje.as_string())
	except Exception as e:
		print(e)	
	
	return redirect('ventas:index')



class create_marketin(ListView):

	model= Producto
	template_name='usuario/listproduc_marketin.html'
	queryset = Producto.objects.all()
	context_object_name ='Producto'
	



	def get_context_data(self , **kwargs):

		context = super().get_context_data(**kwargs)
		context['form']= campana
		
		return context


	def post(self , request ,  *args , **kwargs):

		form  = campana(request.POST)
		print(form)

		

		#import pdb;pdb.set_trace()
		

		if form.is_valid():
			data= form.cleaned_data


			

			
			
		
			valor_add2 =int(data['nombre_campana'])
			#mport pdb;pdb.set_trace()
			#dato = date_capana.objects.create(nombre_campana= data['nombre_campana'])
			dato = date_capana.objects.create(nombre_campana= valor_add2)

			
			dato.producto.set(data['producto'])

			RECIPIENT_LIST =[]
			if data['nombre_campana'] != 1:


				prueba_consulta = Profile.objects.filter(interest=data['nombre_campana'] )
				print(prueba_consulta)
				
				for q in prueba_consulta:
					valor_add = q.user.email
					RECIPIENT_LIST.append(valor_add)
					
				

			else:


				prueba_consulta = Profile.objects.all()
				print(prueba_consulta)
				
				for q in prueba_consulta:
					valor_add = q.user.email
					RECIPIENT_LIST.append(valor_add)
					
				





			
		try:
			mailseerver= smtplib.SMTP(settings.EMAIL_HOST , settings.EMAIL_PORT)
			mailseerver.starttls()
			mailseerver.ehlo()

			mailseerver.login(settings.EMAIL_HOST_USER , settings.EMAIL_HOST_PASSWORD)


			mensaje = MIMEMultipart()
			mensaje['From'] = settings.EMAIL_HOST_USER
			mensaje['To'] = 'coroemma@hotmail.com'
			mensaje['Subject']='tienes un correo'

			dato = date_capana.objects.all()
			longitud = len(dato)
			ultimo = dato[longitud-1]
			envio = ultimo
			


			content = render_to_string('sesion/email3.html' , {'datos': envio})
			mensaje.attach(MIMEText(content , 'html'))
			mailseerver.sendmail(settings.EMAIL_HOST_USER , RECIPIENT_LIST, mensaje.as_string())
		except Exception as e:
			print(e)	
	
		return redirect('ventas:index')

		


		return render(request , 'usuario/listproduc.html')















	



	




